import UI_Constants
import pytesseract
import openpyxl
import pandas as pd
from Mapping_loader import MappingFileLoader


def retrive_dict_value(dictData, keyName, keyParam1):
    dictValue = dictData.get(keyName)
    print("dictValue :", dictValue)
    # print("len(dictValue) :" ,len(dictValue))
    if dictValue == None:
        dictValue = dictData.get(str(keyName) + "_" + str(keyParam1))
        print("str(keyName)+str(keyParam1) :", str(keyName) + "_" + str(keyParam1))

    return dictValue


try:
    # ***********  Reading Mapping file - Product mapping and Parse Param mapping  ****************#
    mf1 = MappingFileLoader.fetch_Parse_param_mapping()
    mf2 = MappingFileLoader.fetch_Product_Mapping()
    mf3 = MappingFileLoader.user_Credentials()
    count_mf2 = len(mf2)
    count_mf3 = len(mf3)

    user_name_with_key = mf3.__getitem__(0)
    pass_word_with_key = mf3.__getitem__(1)

    user_name = user_name_with_key.get('User_Name')
    pass_word = pass_word_with_key.get('Pass_word')
    # print(user_name,pass_word)

    name_retail = mf1.__getitem__(0)
    date_of_sale = mf1.__getitem__(1)
    existing_file_record_range = mf1.__getitem__(2)
    new_file_record_range = mf1.__getitem__(3)
    purchase_record_range = mf1.__getitem__(4)

    name_retail_rows = name_retail.get('name.retail')
    date_of_sale_rows = date_of_sale.get('date.of.sale')
    existing_file_record_range_rows = existing_file_record_range.get('existing.file.record.range')
    new_file_record_range_rows = new_file_record_range.get('new.file.record.range')
    purchase_record_range_rows = purchase_record_range.get('purchase.record.range')

    name_retail__st_row = name_retail_rows.split(',')[0]
    name_retail__end_col = name_retail_rows.split(',')[1]
    date_of_sale__st_row = date_of_sale_rows.split(',')[0]
    date_of_sale__end_col = date_of_sale_rows.split(',')[1]
    existing_file_record_range__st_row = existing_file_record_range_rows.split(',')[0]
    existing_file_record_range__end_row = existing_file_record_range_rows.split(',')[1]
    new_file_record_range__st_row = new_file_record_range_rows.split(',')[0]
    new_file_record_range__end_row = new_file_record_range_rows.split(',')[1]
    purchase_record_range__st_row = purchase_record_range_rows.split(',')[0]
    purchase_record_range__end_row = purchase_record_range_rows.split(',')[1]

    # ************     Reading Main excel sheet    ***************#
    ExcelFile_Path = r'config\SMWSED.xlsx'
    Data_file = openpyxl.load_workbook(ExcelFile_Path)
    Data_sheet = Data_file['SMWSED']

    # ************     Reading Date from excel sheet    ***************#
    Retailer_name = Data_sheet.cell(row=2, column=2).value
    Date_of_sale = Data_sheet.cell(row=3, column=2).value
    Date_of_sale_str = Date_of_sale.split(':')[0]
    Date_of_sale_date = Date_of_sale.split(':')[1]

    # ************    Creating three lists    ***************#
    existing_file_list = []
    new_file_record_list = []
    purchase_record_list = []

    # ************    iterate through excel    ***************#
    for i in range(int(existing_file_record_range__st_row), int(existing_file_record_range__end_row) + 1):
        existing_file_list_col = []
        for j in range(1, Data_sheet.max_column + 1):
            cell_obj = Data_sheet.cell(row=i, column=j)
            existing_file_list_col.append(cell_obj.value)
        existing_file_list.append(existing_file_list_col)

    for i in range(int(new_file_record_range__st_row), int(new_file_record_range__end_row) + 1):
        new_file_record_list_col = []
        for j in range(1, Data_sheet.max_column + 1):
            cell_obj = Data_sheet.cell(row=i, column=j)
            new_file_record_list_col.append(cell_obj.value)
        new_file_record_list.append(new_file_record_list_col)

    for i in range(int(purchase_record_range__st_row), int(purchase_record_range__end_row) + 1):
        purchase_record_list_col = []
        for j in range(1, Data_sheet.max_column + 1):
            cell_obj = Data_sheet.cell(row=i, column=j)
            purchase_record_list_col.append(cell_obj.value)
        purchase_record_list.append(purchase_record_list_col)

    # ************    Getting lenth os three lists    ***************#
    count_pur = len(purchase_record_list)
    count_ex = len(existing_file_list)
    count_new = len(new_file_record_list)

    # ************    creating three lists with PackSize   ***************#
    brandName_packsize_ex = []
    brandName_packsize_new = []
    brandName_packsize_pur = []

    # ************    iterate through excel     ***************#
    for i in range(int(existing_file_record_range__st_row), int(existing_file_record_range__end_row) + 1):
        if not (Data_sheet.cell(row=i, column=2).value == None) or not (Data_sheet.cell(row=i, column=3).value == None):
            brand_name = Data_sheet.cell(row=i, column=2).value
            pack_size = Data_sheet.cell(row=i, column=3).value
            brand_pack_ex = brand_name + "_" + pack_size
            brandName_packsize_ex.append(brand_pack_ex)

    for i in range(int(new_file_record_range__st_row), int(new_file_record_range__end_row) + 1):
        if not (Data_sheet.cell(row=i, column=2).value == None) or not (Data_sheet.cell(row=i, column=3).value == None):
            brand_name = str(Data_sheet.cell(row=i, column=2).value)
            pack_size = str(Data_sheet.cell(row=i, column=3).value)
            brand_pack_new = brand_name + "_" + pack_size
            brandName_packsize_new.append(brand_pack_new)

    for i in range(int(purchase_record_range__st_row), int(purchase_record_range__end_row) + 1):
        if not (Data_sheet.cell(row=i, column=2).value == None) or not (Data_sheet.cell(row=i, column=3).value == None):
            brand_name = Data_sheet.cell(row=i, column=2).value
            pack_size = Data_sheet.cell(row=i, column=3).value
            brand_pack_pur = brand_name + "_" + pack_size
            brandName_packsize_pur.append(brand_pack_pur)

    # ************    Creating class UI ACTIONS    ***************#

    print("**************************Starting  here*************************************")
    #print(mf2)
    print("***************************************************************")


except ImportError:
    print(ImportError)


def fillBrand_Pack_Bottle_Size(salesRecords, purchanseRecords):
    for x in range(0, len(salesRecords)):
        Brand_name_ex = salesRecords[x][1]
        Pack_Size_ex = salesRecords[x][2]
        Bottle_count_ex = salesRecords[x][3]
        is_Exist_Purchase = False

        print("STARTING TO PROCESS ---------------------------------------------------  ",Brand_name_ex)
        print()

        # loop over existing sales
        for y in range(0, len(purchanseRecords)):
            Brand_name_pur = purchanseRecords[y][1]
            Pack_Size_pur = purchanseRecords[y][2]
            Bottle_count_pur = purchanseRecords[y][3]
            # print("Brand_name_pur :",Brand_name_pur, "Brand_name_new:",Brand_name_new )
            if not Brand_name_ex == None:
                is_Exist_Purchase = (Brand_name_ex == Brand_name_pur and Pack_Size_ex == Pack_Size_pur)
                if is_Exist_Purchase:
                    print("****************  BRAND NAME MATCHES WITH PURCHASE - SET VALUE in PURCHASE SECTION *************************")
                    print("Set Value in website ERP Product Name -", Brand_name_ex)
                    print("pack size -", Pack_Size_ex)
                    print("bottle count total -", Bottle_count_ex)
                    print("bottle total in purchase -", Bottle_count_pur)
                    print("Website Product Name    ", retrive_dict_value(mf2, Brand_name_ex, Pack_Size_ex))
                    break
        if is_Exist_Purchase == False:
            print(
                "****************  BRAND NAME **DONOT** MATCHES WITH PURCHASE - SET VALUE in PURCHASE SECTION *************************")
            print("ERP Product Name -", Brand_name_ex)
            print("pack size -", Pack_Size_ex)
            print("bottel count total -", Bottle_count_ex)
            print("Website Product Name    ", retrive_dict_value(mf2, Brand_name_ex, Pack_Size_ex))

# lopp over existing sales
fillBrand_Pack_Bottle_Size(existing_file_list,purchase_record_list)
# lopp over new sales
print("#####################################################################################################")
fillBrand_Pack_Bottle_Size(new_file_record_list,purchase_record_list)
