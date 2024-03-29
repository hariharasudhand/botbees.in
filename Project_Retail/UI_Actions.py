import tagui as r
import UI_Constants
import pytesseract
import openpyxl
import pandas as pd
from Mapping_loader import MappingFileLoader


try:
    def retrive_dict_value(dictData, keyName, keyParam1):
        dictValue = dictData.get(keyName)
        if dictValue == None:
            dictValue = dictData.get(str(keyName) + "_" + str(keyParam1))
        return dictValue


    def setNewAndPurchaseStockValue(brandName, packSize, noofBottle, noofBottlePur, isNewRecord):
        r.wait(2)
        r.click(UI_Constants.XPATH_BRAND_NAME)
        r.wait(1)
        r.type(UI_Constants.XPATH_BRAND_NAME, brandName)
        r.wait(1)
        r.click(UI_Constants.XPATH_CLICK)
        r.wait(1)
        r.click(UI_Constants.XPATH_PACK_SIZE)
        r.wait(1)
        r.type(UI_Constants.XPATH_PACK_SIZE, packSize)
        r.wait(1)
        r.click(UI_Constants.XPATH_CLICK)
        r.wait(1)
        if isNewRecord == True:
            print('')
            r.type(UI_Constants.XPATH_NEW_STOCK, '[clear]')
            r.wait(1)
            r.type(UI_Constants.XPATH_NEW_STOCK, noofBottle)
            r.wait(1)
            r.click(UI_Constants.XPATH_CLICK)
            r.wait(1)
        else:
            print('')
            r.type(UI_Constants.XPATH_EXISTING_STOCK, '[clear]')
            r.wait(1)
            r.type(UI_Constants.XPATH_EXISTING_STOCK, noofBottle)
            r.wait(1)
            r.click(UI_Constants.XPATH_CLICK)
            r.wait(1)
        r.type(UI_Constants.XPATH_NEW_STOCK_PURCHASED, '[clear]')
        r.wait(1)
        r.type(UI_Constants.XPATH_NEW_STOCK_PURCHASED, noofBottlePur)
        r.wait(1)
        r.click(UI_Constants.XPATH_CLICK)
        r.wait(1)
        r.snap('page','result.png')
        # r.click(UI_Constants.XPATH_SUBMIT)
        # r.wait(2)
        # r.click(UI_Constants.IMG_OK_BTN)
        # r.wait(2)
        r.click(UI_Constants.XPATH_BRAND_NAME)
        r.wait(1)
        r.type(UI_Constants.XPATH_BRAND_NAME,UI_Constants.SELECT)
        r.wait(1)
        r.click(UI_Constants.XPATH_CLICK)
        r.wait(1)


    def setExistingStockValue(brandName, packSize, noofBottle, isNewRecord):
        r.wait(2)
        r.click(UI_Constants.XPATH_BRAND_NAME)
        r.wait(1)
        r.type(UI_Constants.XPATH_BRAND_NAME, brandName)
        r.wait(1)
        r.click(UI_Constants.XPATH_CLICK)
        r.wait(1)
        r.click(UI_Constants.XPATH_PACK_SIZE)
        r.wait(1)
        r.type(UI_Constants.XPATH_PACK_SIZE, packSize)
        r.wait(1)
        r.click(UI_Constants.XPATH_CLICK)
        r.wait(1)
        if isNewRecord == 1:
            print("")
            r.type(UI_Constants.XPATH_NEW_STOCK, '[clear]')
            r.wait(1)
            r.type(UI_Constants.XPATH_NEW_STOCK, noofBottle)
            r.wait(1)
            r.click(UI_Constants.XPATH_CLICK)
            r.wait(1)
        elif isNewRecord == 2:
            print("")
            r.type(UI_Constants.XPATH_EXISTING_STOCK, '[clear]')
            r.wait(1)
            r.type(UI_Constants.XPATH_EXISTING_STOCK, noofBottle)
            r.wait(1)
            r.click(UI_Constants.XPATH_CLICK)
            r.wait(1)
        elif isNewRecord == 3:
            print("")
            r.type(UI_Constants.XPATH_NEW_STOCK_PURCHASED, '[clear]')
            r.wait(1)
            r.type(UI_Constants.XPATH_NEW_STOCK_PURCHASED, noofBottle)
            r.wait(1)
            r.click(UI_Constants.XPATH_CLICK)
            r.wait(1)
        r.snap('page', 'result.png')
        # r.click(UI_Constants.XPATH_SUBMIT)
        # r.wait(2)
        # r.click(UI_Constants.IMG_OK_BTN)
        # r.wait(2)
        r.click(UI_Constants.XPATH_BRAND_NAME)
        r.wait(1)
        r.type(UI_Constants.XPATH_BRAND_NAME, UI_Constants.SELECT)
        r.wait(1)
        r.click(UI_Constants.XPATH_CLICK)
        r.wait(1)


    # ***********  Reading Mapping file - Product mapping and Parse Param mapping  ****************#
    mf1 = MappingFileLoader.fetch_Parse_param_mapping()
    mf2 = MappingFileLoader.fetch_Product_Mapping()
    mf3 = MappingFileLoader.user_Credentials()
    username = mf3.get('User_Name')
    password = mf3.get('Pass_word')
    login_retry = mf3.get('LOGIN_RETRY')
    count_mf2 = len(mf2)

    name_retail = mf1.__getitem__(0)
    date_of_sale = mf1.__getitem__(1)
    existing_file_record_range = mf1.__getitem__(2)
    new_file_record_range = mf1.__getitem__(3)
    purchase_record_range = mf1.__getitem__(4)

    name_retail_rows = name_retail.get('name.retail')
    date_of_sale_rows = date_of_sale.get('date.of.sale')
    existing_file_record_range_rows = existing_file_record_range.get('existing.file.record.range')
    new_file_record_range_rows = new_file_record_range.get('new.file.record.range')
    purchase_record_range_rows = purchase_record_range.get('purchase.record.range')

    name_retail__st_row = name_retail_rows.split(',')[0]
    name_retail__end_col = name_retail_rows.split(',')[1]
    date_of_sale__st_row = date_of_sale_rows.split(',')[0]
    date_of_sale__end_col = date_of_sale_rows.split(',')[1]
    existing_file_record_range__st_row = existing_file_record_range_rows.split(',')[0]
    existing_file_record_range__end_row = existing_file_record_range_rows.split(',')[1]
    new_file_record_range__st_row = new_file_record_range_rows.split(',')[0]
    new_file_record_range__end_row = new_file_record_range_rows.split(',')[1]
    purchase_record_range__st_row = purchase_record_range_rows.split(',')[0]
    purchase_record_range__end_row = purchase_record_range_rows.split(',')[1]




    # ************     Reading Main excel sheet    ***************#
    ExcelFile_Path = r'C:/Users/Admin/.botbees/config/SMWSED.xlsx'
    Data_file = openpyxl.load_workbook(ExcelFile_Path)
    Data_sheet = Data_file['SMWSED']

    # ************     Reading Date from excel sheet    ***************#
    Retailer_name = Data_sheet.cell(row=2, column=2).value
    Date_of_sale = Data_sheet.cell(row=3, column=2).value
    Date_of_sale_str = Date_of_sale.split(':')[0]
    Date_of_sale_date = Date_of_sale.split(':')[1]

    # ************    Creating three lists    ***************#
    existing_file_list = []
    new_file_record_list = []
    purchase_record_list = []



    # ************    iterate through excel    ***************#
    for i in range(int(existing_file_record_range__st_row), int(existing_file_record_range__end_row) + 1):
        existing_file_list_col = []
        for j in range(1, Data_sheet.max_column + 1):
            cell_obj = Data_sheet.cell(row=i, column=j)
            existing_file_list_col.append(cell_obj.value)
        existing_file_list.append(existing_file_list_col)


    for i in range(int(new_file_record_range__st_row), int(new_file_record_range__end_row) + 1):
        new_file_record_list_col = []
        for j in range(1, Data_sheet.max_column + 1):
            cell_obj = Data_sheet.cell(row=i, column=j)
            new_file_record_list_col.append(cell_obj.value)
        new_file_record_list.append(new_file_record_list_col)



    for i in range(int(purchase_record_range__st_row), int(purchase_record_range__end_row) + 1):
        purchase_record_list_col = []
        for j in range(1, Data_sheet.max_column + 1):
            cell_obj = Data_sheet.cell(row=i, column=j)
            purchase_record_list_col.append(cell_obj.value)
        purchase_record_list.append(purchase_record_list_col)


    # ************    Getting lenth os three lists    ***************#
    count_pur = len(purchase_record_list)
    count_ex = len(existing_file_list)
    count_new = len(new_file_record_list)

    # ************    Creating class UI ACTIONS    ***************#

    class UI_Actions:
        def __init__(self):
            print()
            r.init(visual_automation=True)
            r.url(UI_Constants.LOGIN_URL)
            r.wait(5)

        def login(self):
            print("its inside login method")
            if r.popup(UI_Constants.LOGIN_URL) == True:
                r.wait(5)
                r.type(UI_Constants.XPATH_USERNAME, username)
                r.wait(2)
                r.type(UI_Constants.XPATH_PASSWORD,password)
                r.wait(2)
                r.snap(UI_Constants.XPATH_CAPTCHA_IMG, UI_Constants.IMG_CAPTCHA_CAPTUREIMAGE_NAME)
                TextFromImage = pytesseract.image_to_string(UI_Constants.IMG_CAPTCHA_CAPTUREIMAGE_NAME)
                r.wait(5)
                r.type(UI_Constants.XPATH_CAPTCHA_TEXT,TextFromImage)
                # r.click(UI_Constants.XPATH_LOGIN_BTN)
                if r.exist(UI_Constants.XPATH_LOGIN_BTN) == True:
                    # r.click(UI_Constants.ok_login)
                    r.wait(3)
                    r.click(UI_Constants.ok_login)
                    r.wait(10)
                    r.click(UI_Constants.ok_login)
                r.wait(5)
                isLoginSuccess = (r.click(UI_Constants.XPATH_LOGIN_BTN) == False)
                return isLoginSuccess

        def manualLogin(self):
            user_name = r.ask("Enter User Name ")
            r.wait(10)
            r.type(UI_Constants.XPATH_USERNAME, user_name)
            r.wait(10)
            pass_word = r.ask("Enter Pass word ")
            r.wait(10)
            r.type(UI_Constants.XPATH_PASSWORD, pass_word)
            r.wait(10)
            captcha_text = r.ask("Enter Captcha Text ")
            r.wait(10)
            r.type(UI_Constants.XPATH_CAPTCHA_TEXT, captcha_text)
            r.wait(10)
            r.click(UI_Constants.XPATH_LOGIN_BTN)
            r.wait(1)

        def performAction(self):
            r.wait(1)
            r.click(UI_Constants.XPATH_PERMITS_LINK)
            r.wait(1)
            r.click(UI_Constants.XPATH_DEALER_LINK)

        def dateAndFL2reatailSelection(self):
            r.wait(2)
            r.type(UI_Constants.XPATH_DATE_OF_SALE, Date_of_sale_date)
            r.wait(2)
            r.type(UI_Constants.XPATH_FL2_REATAIL_SALE, UI_Constants.FL2_REATAIL_SALE)
            r.wait(1)


        @staticmethod
        def fillingBrandnamePacksizeAndBottles(salesRecords, purchanseRecords):
            for x in range(0, len(salesRecords)):
                Brand_name_ex = str(salesRecords[x][1])
                Pack_Size_ex = str(salesRecords[x][2])
                Bottle_count_ex = str(salesRecords[x][3])
                is_Exist_Purchase = False
                # loop over existing sales
                for y in range(0, len(purchanseRecords)):
                    Brand_name_pur = str(purchanseRecords[y][1])
                    Pack_Size_pur = str(purchanseRecords[y][2])
                    Bottle_count_pur = str(purchanseRecords[y][3])
                    if not Brand_name_ex == None:
                        is_Exist_Purchase = (Brand_name_ex == Brand_name_pur and Pack_Size_ex == Pack_Size_pur)
                        if is_Exist_Purchase:
                            if salesRecords == existing_file_list:
                                setNewAndPurchaseStockValue(retrive_dict_value(mf2, Brand_name_ex, Pack_Size_ex),Pack_Size_ex, Bottle_count_ex, Bottle_count_pur, False)
                            elif salesRecords == new_file_record_list:
                                setNewAndPurchaseStockValue(retrive_dict_value(mf2, Brand_name_ex, Pack_Size_ex),Pack_Size_ex, Bottle_count_ex, Bottle_count_pur, True)
                            break
                if is_Exist_Purchase == False:
                    if salesRecords == purchase_record_list:
                        setExistingStockValue(retrive_dict_value(mf2, Brand_name_ex, Pack_Size_ex), Pack_Size_ex,Bottle_count_ex, 3)
                    elif salesRecords == existing_file_list:
                        if salesRecords[x][1] == 0 :
                            break
                        else:
                            setExistingStockValue(retrive_dict_value(mf2, Brand_name_ex, Pack_Size_ex), Pack_Size_ex, Bottle_count_ex, 2)
                    elif salesRecords == new_file_record_list:
                        setExistingStockValue(retrive_dict_value(mf2, Brand_name_ex, Pack_Size_ex), Pack_Size_ex,Bottle_count_ex, 1)


        def calling_fill(self):
            self.fillingBrandnamePacksizeAndBottles(existing_file_list, purchase_record_list)
            self.fillingBrandnamePacksizeAndBottles(new_file_record_list, purchase_record_list)
            self.fillingBrandnamePacksizeAndBottles(purchase_record_list, new_file_record_list)

        def logOut(self):
            r.wait(2)
            r.click(UI_Constants.XPATH_LOG_OUT)
            r.wait(2)
            r.close()

        def close(self):
            r.close()

except ImportError:
    print(ImportError)

